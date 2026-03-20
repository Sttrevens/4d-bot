"""文件导出工具 —— 生成文件并通过企微/飞书发送给用户

支持格式: CSV, TXT, Markdown, JSON, PDF, HTML, XLSX
适用平台: wecom, wecom_kf（飞书用 doc_ops 创建云文档，不需要文件导出）

工作流程:
1. LLM 调用 export_file 工具，传入文件名 + 内容
2. 工具生成文件字节
   - PDF: LLM 生成 HTML+CSS → weasyprint 渲染为 PDF（支持完整排版/链接/图片/表格样式）
   - 其他格式: 直接编码
3. 通过企微临时素材 API 上传
4. 发送文件消息给当前用户

企微临时素材限制:
- 文件上限 20MB
- 有效期 3 天
- 支持类型: image/voice/video/file
"""

from __future__ import annotations

import csv
import io
import json
import logging
import os
import re
import time

import httpx

from app.tools.tool_result import ToolResult
from app.tenant.context import get_current_tenant
from app.tools.feishu_api import _current_user_open_id
from app.tools.source_registry import sanitize_urls_in_content

logger = logging.getLogger(__name__)

# 抑制 fontTools 在 PDF 字体子集化时的海量 DEBUG/INFO 日志
logging.getLogger("fontTools.subset").setLevel(logging.WARNING)

# ── 企微 API（同步，供工具函数调用）──

_TOKEN_URL = "https://qyapi.weixin.qq.com/cgi-bin/gettoken"
_UPLOAD_URL = "https://qyapi.weixin.qq.com/cgi-bin/media/upload"
_SEND_URL = "https://qyapi.weixin.qq.com/cgi-bin/message/send"        # 企微内部应用
_KF_SEND_URL = "https://qyapi.weixin.qq.com/cgi-bin/kf/send_msg"     # 微信客服

# token 缓存: {cache_key: (token, expire_time)}
_token_cache: dict[str, tuple[str, float]] = {}

# 文件大小上限 (20MB)
_MAX_FILE_SIZE = 20 * 1024 * 1024


def _get_token_sync(corpid: str, secret: str, cache_key: str = "") -> str:
    """同步获取 access_token（带缓存）"""
    ck = cache_key or corpid
    cached = _token_cache.get(ck)
    if cached and time.time() < cached[1]:
        return cached[0]

    with httpx.Client(timeout=10, trust_env=False) as client:
        resp = client.get(_TOKEN_URL, params={"corpid": corpid, "corpsecret": secret})
        data = resp.json()

    if data.get("errcode", -1) != 0:
        raise RuntimeError(f"wecom token error: {data}")

    token = data["access_token"]
    expire = time.time() + data.get("expires_in", 7200) - 300
    _token_cache[ck] = (token, expire)
    return token


def _upload_media_sync(token: str, file_bytes: bytes, filename: str) -> str:
    """同步上传临时素材，返回 media_id"""
    with httpx.Client(timeout=30, trust_env=False) as client:
        resp = client.post(
            _UPLOAD_URL,
            params={"access_token": token, "type": "file"},
            # httpx files 需要 bytes（不是 bytearray），用 io.BytesIO 包装最安全
            files={"media": (filename, io.BytesIO(file_bytes))},
        )
        data = resp.json()

    if data.get("errcode", 0) != 0:
        raise RuntimeError(f"upload failed: {data}")

    return data.get("media_id", "")


def _send_file_wecom(token: str, userid: str, media_id: str, agent_id: int) -> dict:
    """企微内部应用发送文件"""
    body = {
        "touser": userid,
        "msgtype": "file",
        "agentid": agent_id,
        "file": {"media_id": media_id},
    }
    with httpx.Client(timeout=10, trust_env=False) as client:
        resp = client.post(_SEND_URL, params={"access_token": token}, json=body)
        return resp.json()


def _send_file_wecom_kf(token: str, external_userid: str, media_id: str, open_kfid: str) -> dict:
    """微信客服发送文件"""
    body = {
        "touser": external_userid,
        "open_kfid": open_kfid,
        "msgtype": "file",
        "file": {"media_id": media_id},
    }
    with httpx.Client(timeout=10, trust_env=False) as client:
        resp = client.post(_KF_SEND_URL, params={"access_token": token}, json=body)
        return resp.json()


# ── 文件生成 ──

def _generate_csv(content: str) -> bytes:
    """从结构化文本生成 CSV 文件字节

    content 格式: 每行一条记录，字段用逗号分隔。第一行为表头。
    也支持传入 JSON 数组格式（自动转 CSV）。
    """
    content = content.strip()

    # 尝试解析 JSON 数组
    if content.startswith("["):
        try:
            rows = json.loads(content)
            if rows and isinstance(rows[0], dict):
                buf = io.StringIO()
                writer = csv.DictWriter(buf, fieldnames=rows[0].keys())
                writer.writeheader()
                writer.writerows(rows)
                return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
        except (json.JSONDecodeError, TypeError, AttributeError):
            pass

    # 原样文本（已经是 CSV 格式）
    return content.encode("utf-8-sig")


def _generate_json(content: str) -> bytes:
    """格式化 JSON 输出"""
    try:
        parsed = json.loads(content)
        return json.dumps(parsed, ensure_ascii=False, indent=2).encode("utf-8")
    except json.JSONDecodeError:
        return content.encode("utf-8")


def _generate_xlsx(content: str) -> bytes:
    """从结构化文本生成 Excel (.xlsx) 文件

    content 格式:
    - JSON 数组（推荐）: [{"列A": "值1", "列B": "值2"}, ...]
    - CSV 文本: 每行逗号分隔，第一行为表头
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active

    content = content.strip()
    rows: list[list[str]] = []
    headers: list[str] = []

    # 尝试解析 JSON 数组
    if content.startswith("["):
        try:
            data = json.loads(content)
            if data and isinstance(data[0], dict):
                headers = list(data[0].keys())
                rows = [[str(row.get(h, "")) for h in headers] for row in data]
        except (json.JSONDecodeError, TypeError, AttributeError, IndexError):
            pass

    # 回退: 按 CSV 解析
    if not rows:
        reader = csv.reader(io.StringIO(content))
        all_rows = list(reader)
        if all_rows:
            headers = all_rows[0]
            rows = all_rows[1:]

    if not headers:
        # 纯文本兜底: 每行一个单元格
        for i, line in enumerate(content.split("\n"), 1):
            ws.cell(row=i, column=1, value=line)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 写表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 写数据行（偶数行浅灰底色）
    even_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = even_fill

    # 自动列宽（取表头和前 50 行数据的最大宽度）
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_data in rows[:50]:
            if col_idx - 1 < len(row_data):
                max_len = max(max_len, len(str(row_data[col_idx - 1])))
        # CJK 字符宽度约 2x，简单估算
        adjusted = min(max_len + 4, 60)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = adjusted

    # 冻结首行
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF 生成（weasyprint HTML→PDF 管线）──

# weasyprint 基础 CSS：注入 CJK 字体 + 合理的页面默认样式
# LLM 生成的 HTML 会被包裹在这个基础样式之上
_BASE_PDF_CSS = """
@page {
    size: A4;
    margin: 2cm 2.5cm;
}
body {
    font-family: 'Noto Sans SC', 'PingFang SC', 'Microsoft YaHei', sans-serif;
    font-size: 11pt;
    line-height: 1.6;
    color: #333;
}
h1 { font-size: 22pt; font-weight: 700; margin: 0.5em 0 0.3em; color: #111; }
h2 { font-size: 16pt; font-weight: 700; margin: 0.8em 0 0.3em; color: #222; border-bottom: 1px solid #ddd; padding-bottom: 0.2em; }
h3 { font-size: 13pt; font-weight: 700; margin: 0.6em 0 0.2em; color: #333; }
table { width: 100%; border-collapse: collapse; margin: 0.8em 0; font-size: 10pt; }
th { background: #f5f5f5; font-weight: 700; text-align: left; padding: 6px 8px; border: 1px solid #ddd; }
td { padding: 6px 8px; border: 1px solid #ddd; }
tr:nth-child(even) { background: #fafafa; }
a { color: #0066cc; text-decoration: underline; }
ul, ol { padding-left: 1.5em; }
li { margin: 0.2em 0; }
blockquote { border-left: 3px solid #ddd; padding: 0.3em 1em; margin: 0.5em 0; color: #666; background: #f9f9f9; }
hr { border: none; border-top: 1px solid #ddd; margin: 1em 0; }
code { background: #f4f4f4; padding: 1px 4px; border-radius: 3px; font-size: 0.9em; }
pre { background: #f4f4f4; padding: 0.8em; border-radius: 4px; overflow-x: auto; font-size: 9pt; }
img { max-width: 100%; height: auto; }
.page-break { page-break-after: always; }
"""


def _is_html_content(content: str) -> bool:
    """检测内容是否为 HTML。"""
    s = content.strip()[:200].lower()
    return s.startswith(("<!doctype", "<html", "<head", "<body", "<div", "<style",
                         "<table", "<section", "<article", "<main"))


def _html_to_pdf(html: str) -> bytes | None:
    """用 weasyprint 将 HTML 转为 PDF。返回 None 表示 weasyprint 不可用。"""
    try:
        from weasyprint import HTML, CSS
    except ImportError:
        logger.info("weasyprint not installed, falling back to fpdf2")
        return None

    try:
        # 注入基础 CSS（LLM 的内联 CSS 优先级更高，会覆盖基础样式）
        result = HTML(string=html).write_pdf(
            stylesheets=[CSS(string=_BASE_PDF_CSS)]
        )
        logger.info("weasyprint: HTML→PDF OK (%dKB)", len(result) // 1024)
        return bytes(result)
    except Exception:
        logger.warning("weasyprint render failed", exc_info=True)
        return None


def _markdown_to_minimal_html(md: str) -> str:
    """将 Markdown 文本转为最基础的 HTML（无需外部库）。
    用于 weasyprint 渲染 Markdown 内容（比 fpdf2 效果好得多）。
    """
    lines = md.split("\n")
    html_parts = ["<html><body>"]
    in_table = False
    in_ul = False
    in_ol = False

    def _inline(text: str) -> str:
        """处理行内格式：**bold**, [link](url)"""
        # bold
        text = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", text)
        # links
        text = re.sub(r"\[([^\]]+)\]\(([^)]+)\)", r'<a href="\2">\1</a>', text)
        # inline code
        text = re.sub(r"`([^`]+)`", r"<code>\1</code>", text)
        return text

    def _close_list():
        nonlocal in_ul, in_ol
        if in_ul:
            html_parts.append("</ul>")
            in_ul = False
        if in_ol:
            html_parts.append("</ol>")
            in_ol = False

    for line in lines:
        stripped = line.strip()

        if not stripped:
            if in_table:
                html_parts.append("</tbody></table>")
                in_table = False
            _close_list()
            continue

        # 表格
        if "|" in stripped and stripped.startswith("|"):
            cells = [c.strip() for c in stripped.split("|")[1:-1]]
            # 分隔行（---）跳过
            if all(c.replace("-", "").replace(":", "") == "" for c in cells):
                continue
            if not in_table:
                html_parts.append("<table><thead><tr>")
                for c in cells:
                    html_parts.append(f"<th>{_inline(c)}</th>")
                html_parts.append("</tr></thead><tbody>")
                in_table = True
            else:
                html_parts.append("<tr>")
                for c in cells:
                    html_parts.append(f"<td>{_inline(c)}</td>")
                html_parts.append("</tr>")
            continue

        if in_table:
            html_parts.append("</tbody></table>")
            in_table = False

        # 标题
        if stripped.startswith("### "):
            _close_list()
            html_parts.append(f"<h3>{_inline(stripped[4:])}</h3>")
            continue
        if stripped.startswith("## "):
            _close_list()
            html_parts.append(f"<h2>{_inline(stripped[3:])}</h2>")
            continue
        if stripped.startswith("# "):
            _close_list()
            html_parts.append(f"<h1>{_inline(stripped[2:])}</h1>")
            continue

        # 分隔线
        if stripped.startswith("---") and len(set(stripped)) <= 2:
            _close_list()
            html_parts.append("<hr>")
            continue

        # 无序列表
        if stripped.startswith(("- ", "* ", "• ")):
            if not in_ul:
                _close_list()
                html_parts.append("<ul>")
                in_ul = True
            html_parts.append(f"<li>{_inline(stripped[2:])}</li>")
            continue

        # 有序列表
        m = re.match(r"^(\d+)\.\s+(.+)", stripped)
        if m:
            if not in_ol:
                _close_list()
                html_parts.append("<ol>")
                in_ol = True
            html_parts.append(f"<li>{_inline(m.group(2))}</li>")
            continue

        # 引用
        if stripped.startswith("> "):
            _close_list()
            html_parts.append(f"<blockquote>{_inline(stripped[2:])}</blockquote>")
            continue

        # 普通段落
        _close_list()
        html_parts.append(f"<p>{_inline(stripped)}</p>")

    if in_table:
        html_parts.append("</tbody></table>")
    _close_list()
    html_parts.append("</body></html>")
    return "\n".join(html_parts)


def _generate_pdf(content: str) -> bytes:
    """生成 PDF。自动检测内容格式：

    1. HTML 内容 → weasyprint 直接渲染（完整 CSS 排版/超链接/图片/表格样式）
    2. Markdown 内容 → 转为 HTML → weasyprint 渲染
    3. weasyprint 不可用 → 回退到 fpdf2（基础排版，无图片/链接支持）
    """
    if _is_html_content(content):
        # LLM 生成的 HTML，直接用 weasyprint
        result = _html_to_pdf(content)
        if result:
            return result
        # weasyprint 不可用，保存为纯 HTML（用户至少能在浏览器打开）
        logger.warning("weasyprint unavailable for HTML content, returning raw HTML as PDF fallback")
    else:
        # Markdown 内容 → 转 HTML → weasyprint
        html = _markdown_to_minimal_html(content)
        result = _html_to_pdf(html)
        if result:
            return result

    # 最终兜底：fpdf2（旧路径，能力有限但总比没有好）
    return _generate_pdf_fpdf2(content)


def _generate_pdf_fpdf2(content: str) -> bytes:
    """fpdf2 兜底 PDF 生成（weasyprint 不可用时使用）。
    功能有限：无图片、无可点击链接、CJK 加粗可能不工作。
    """
    from fpdf import FPDF

    pdf = FPDF(format="A4")
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # CJK 字体
    font_path = _find_cjk_font()
    if font_path:
        try:
            pdf.add_font("CJK", fname=font_path)
            font = "CJK"
        except Exception:
            font = "Helvetica"
    else:
        font = "Helvetica"

    pw = pdf.w - pdf.l_margin - pdf.r_margin
    pdf.set_font(font, "", 10)
    pdf.set_text_color(50, 50, 50)

    for line in content.split("\n"):
        stripped = line.strip()
        if not stripped:
            pdf.ln(3)
            continue
        if stripped.startswith("# "):
            pdf.set_font(font, "", 18)
            pdf.multi_cell(pw, 8, stripped[2:])
            pdf.ln(4)
            pdf.set_font(font, "", 10)
        elif stripped.startswith("## "):
            pdf.set_font(font, "", 14)
            pdf.multi_cell(pw, 7, stripped[3:])
            pdf.ln(3)
            pdf.set_font(font, "", 10)
        elif stripped.startswith("### "):
            pdf.set_font(font, "", 12)
            pdf.multi_cell(pw, 6, stripped[4:])
            pdf.ln(2)
            pdf.set_font(font, "", 10)
        elif stripped.startswith(("- ", "* ")):
            pdf.cell(8, 6, chr(8226))
            pdf.multi_cell(pw - 8, 6, stripped[2:])
        else:
            pdf.multi_cell(pw, 6, stripped)
    return bytes(pdf.output())


# CJK 字体（fpdf2 兜底用）
_CJK_FONT_PATHS = [
    "/usr/share/fonts/truetype/noto/NotoSansSC.ttf",
    "/tmp/NotoSansSC.ttf",
]
_FONT_DOWNLOAD_URLS = [
    "https://cdn.jsdelivr.net/gh/google/fonts@main/ofl/notosanssc/NotoSansSC%5Bwght%5D.ttf",
    "https://raw.githubusercontent.com/google/fonts/main/ofl/notosanssc/NotoSansSC%5Bwght%5D.ttf",
]


def _find_cjk_font() -> str | None:
    for p in _CJK_FONT_PATHS:
        if os.path.exists(p) and os.path.getsize(p) > 1_000_000:
            return p
    # 运行时下载兜底
    import urllib.request
    out = "/tmp/NotoSansSC.ttf"
    for url in _FONT_DOWNLOAD_URLS:
        try:
            urllib.request.urlretrieve(url, out)
            if os.path.getsize(out) > 1_000_000:
                return out
        except Exception:
            pass
    return None


# ── 工具定义 ──

TOOL_DEFINITIONS = [
    {
        "name": "export_file",
        "description": (
            "生成文件并发送给用户。适用于用户要求产出报告、表格、数据导出、演示文稿等场景。"
            "支持格式: csv, txt, md, json, pdf, html, xlsx。"
            "XLSX（Excel）推荐用于结构化数据——报价单、设备清单、巡检表等，用户可直接在手机/电脑打开编辑。"
            "传入 JSON 数组（推荐）或 CSV 文本，自动生成带表头样式、斑马纹、冻结首行的专业 Excel 表格。"
            "PDF 推荐用于正式报告——传入完整 HTML+CSS 内容，自动渲染为高质量 PDF"
            "（支持粗体/颜色/表格样式/超链接/base64图片/分页等，远比 Markdown 强大）。"
            "也支持传入 Markdown 文本（自动转为 HTML 再渲染 PDF）。"
            "HTML 可用于演示文稿/PPT/slides——生成单文件 HTML 演示文稿（每页 100vh section，scroll-snap 导航）。"
            "用户要PPT/幻灯片时必须用 .html 格式，不要用 .md！"
            "重要：报告/PDF 中的超链接必须来自 web_search 返回的真实 URL，禁止编造任何链接。"
            "文件会作为聊天附件发送，用户可以直接下载。"
            "注意: 飞书平台请使用 create_feishu_doc 创建云文档，不要用此工具。"
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "filename": {
                    "type": "string",
                    "description": "文件名（含扩展名），如 report.pdf, data.csv, summary.md",
                },
                "content": {
                    "type": "string",
                    "description": (
                        "文件内容。"
                        "XLSX: 推荐传 JSON 数组 [{\"列名\": \"值\"}, ...]，也可传 CSV 文本。"
                        "自动生成带表头样式、斑马纹、冻结首行的专业 Excel 表格。"
                        "PDF: 推荐传完整 HTML（含 <style> 内联 CSS），可用 <strong> 加粗、"
                        "<a href> 超链接、<table> 带样式表格、<img src='data:...'> 内嵌图片、"
                        "CSS color/background 配色等。也可传 Markdown 文本（自动转 HTML）。"
                        "CSV: 每行逗号分隔或 JSON 数组; TXT/MD: 纯文本; JSON: JSON 字符串; "
                        "HTML: 直接保存为 .html 文件"
                    ),
                },
            },
            "required": ["filename", "content"],
        },
    },
]


def _export_file(filename: str, content: str) -> ToolResult:
    """生成文件 → 上传 → 发送给用户"""
    tenant = get_current_tenant()
    platform = tenant.platform
    sender_id = _current_user_open_id.get("")

    if not sender_id:
        return ToolResult.error("无法确定当前用户，无法发送文件", code="internal")

    # 飞书平台应使用云文档
    if platform == "feishu":
        return ToolResult.error(
            "飞书平台请使用 create_feishu_doc 创建云文档，而不是文件导出。",
            code="invalid_param",
            retry_hint="改用 create_feishu_doc 工具创建飞书云文档",
        )

    if platform not in ("wecom", "wecom_kf"):
        return ToolResult.error(
            f"当前平台 {platform} 不支持文件发送", code="invalid_param",
            retry_hint="该平台不支持文件导出，请直接在回复中展示内容",
        )

    # 架构级 URL 验证：在生成文件前，移除未经 web_search 验证的 URL
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "txt"
    if ext in ("pdf", "html"):
        content, removed_urls = sanitize_urls_in_content(content)
        if removed_urls:
            logger.info(
                "export_file: sanitized %d unverified URLs from %s",
                len(removed_urls), filename,
            )

    # 生成文件字节
    if ext == "xlsx":
        try:
            file_bytes = _generate_xlsx(content)
        except Exception:
            logger.exception("XLSX generation failed, falling back to .csv")
            filename = filename.rsplit(".", 1)[0] + ".csv"
            file_bytes = _generate_csv(content)
    elif ext == "pdf":
        try:
            file_bytes = _generate_pdf(content)
        except Exception:
            logger.exception("PDF generation failed, falling back to .html")
            filename = filename.rsplit(".", 1)[0] + ".html"
            file_bytes = content.encode("utf-8")
    elif ext == "csv":
        file_bytes = _generate_csv(content)
    elif ext == "json":
        file_bytes = _generate_json(content)
    else:
        file_bytes = content.encode("utf-8")

    if len(file_bytes) > _MAX_FILE_SIZE:
        return ToolResult.error(
            f"文件太大（{len(file_bytes) // 1024}KB），企微限制 20MB",
            code="invalid_param",
        )

    try:
        # 获取 token
        if platform == "wecom":
            token = _get_token_sync(tenant.wecom_corpid, tenant.wecom_corpsecret)
        else:  # wecom_kf
            token = _get_token_sync(
                tenant.wecom_corpid,
                tenant.wecom_kf_secret,
                cache_key=f"{tenant.wecom_corpid}:kf",
            )

        # 上传临时素材
        media_id = _upload_media_sync(token, file_bytes, filename)
        if not media_id:
            return ToolResult.error("文件上传失败，media_id 为空", code="api_error")

        # 发送文件消息
        if platform == "wecom":
            result = _send_file_wecom(token, sender_id, media_id, tenant.wecom_agent_id)
        else:  # wecom_kf
            result = _send_file_wecom_kf(token, sender_id, media_id, tenant.wecom_kf_open_kfid)

        if result.get("errcode", -1) != 0:
            return ToolResult.error(f"文件发送失败: {result}", code="api_error")

        size_kb = len(file_bytes) / 1024
        return ToolResult.success(
            f"文件 {filename} 已发送给用户（{size_kb:.1f}KB）。"
            f"用户可在聊天中直接下载。"
        )

    except Exception as exc:
        logger.exception("export_file failed: %s", filename)
        return ToolResult.error(f"文件导出失败: {exc}", code="internal")


TOOL_MAP = {
    "export_file": lambda args: _export_file(
        filename=args["filename"],
        content=args["content"],
    ),
}
